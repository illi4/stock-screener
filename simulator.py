# Simulates trading progress and results over time using a spreadsheet with various considerations
# Will save the output result to simulator_result.csv

# Suppress warnings from urllib and gspread
import warnings
warnings.filterwarnings("ignore")

from peewee import IntegrityError

import libs.gsheetobj as gsheetsobj
from libs.signal import red_day_on_volume
from datetime import datetime, timedelta
from libs.stocktools import get_stock_data, Market
import argparse

import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, numbers

# For plotting
import matplotlib.pyplot as plt
import matplotlib.dates as mdates
from openpyxl.drawing.image import Image
import io

parser = argparse.ArgumentParser()

from libs.read_settings import read_config
config = read_config()

from libs.simulation import Simulation
from libs.db import check_earliest_price_date, delete_all_prices, bulk_add_prices, get_price_from_db

pd.set_option("display.max_columns", None)

def create_variant_plot(sim, variant_name):
    # Convert dates to datetime objects
    dates = [datetime.strptime(date, "%d/%m/%Y") for date in sim.detailed_capital_values.keys()]
    values = list(sim.detailed_capital_values.values())

    # Add termination value as a step change
    last_date = dates[-1]
    next_month = (last_date.replace(day=1) + timedelta(days=32)).replace(day=1)

    plt.figure(figsize=(12, 6))

    # Plot the main line
    plt.step(dates, values, where='post')

    # Add the final step
    plt.step([last_date, next_month], [values[-1], sim.current_capital], where='post')

    plt.title(f"Capital Over Time - {variant_name}")
    plt.xlabel("Date")
    plt.ylabel("Capital")

    # Set x-axis to show only the first of each month
    plt.gca().xaxis.set_major_locator(mdates.MonthLocator())
    plt.gca().xaxis.set_major_formatter(mdates.DateFormatter('%Y-%m-%01'))

    # Extend x-axis slightly to show the final step
    plt.xlim(dates[0], next_month + timedelta(days=1))

    plt.xticks(rotation=45)
    plt.tight_layout()

    # Save plot to a bytes buffer
    buf = io.BytesIO()
    plt.savefig(buf, format='png', dpi=300)
    buf.seek(0)
    plt.close()
    return buf

def adjust_column_width(worksheet):
    for col in worksheet.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        worksheet.column_dimensions[column].width = adjusted_width

def set_font_size(worksheet, size):
    for row in worksheet.iter_rows():
        for cell in row:
            cell.font = Font(size=size)

def format_percentage(value):
    return f"{value:.2%}"

def format_number(value):
    return f"{value:.2f}"

def define_args():
    # Take profit levels variation is only supported for the control group, thus the modes are different
    # Removing this as not used now, only one mode using the sheet
    # parser.add_argument(
    #     "-mode",
    #     type=str,
    #     required=True,
    #     help="Mode to run the simulation in (main|tp). Main mode means taking profit as per the spreadsheet setup.",
    #     choices=["main", "tp"],
    # )
    parser.add_argument(
        "--plot", action="store_true", help="Plot the latest simulation"
    )
    parser.add_argument(
        "--failsafe", action="store_true", help="Activate the failsafe approach"
    )
    parser.add_argument(
        "--forced_price_update", action="store_true", help="Activate the failsafe approach"
    )
    #
    # parser.add_argument(
    #     "--show_monthly", action="store_true", help="Show MoM capital value (only in main mode)"
    # )

    parser.add_argument(
        "-method",
        type=str,
        required=True,
        choices=["mri", "anx"],
        help="Method of shortlisting (mri or anx)"
    )

    # Adding the dates
    parser.add_argument(
        "-start",
        type=str,
        required=True,
        help="Start date to run for (YYYY-MM-DD format)",
    )
    parser.add_argument(
        "-end",
        type=str,
        required=True,
        help="End date to run for (YYYY-MM-DD format)",
    )

    # Not used
    # # Arguments to overwrite default settings for filtering
    # parser.add_argument(
    #     "--red_day_exit",
    #     action="store_true",
    #     help="Exit when entry day is red (in tp mode only)",
    # )

    # Not used
    # Test for the exit price variation experimentation
    # parser.add_argument(
    #     "--exit_variation_a",
    #     action="store_true",
    #     help="Exit approach experiment A (main mode only)",
    # )

    args = parser.parse_args()
    arguments = vars(args)

    # Convert specific arguments to boolean, defaulting to False if not provided
    boolean_args = ["plot", "failsafe", "forced_price_update"]   # "show_monthly"
    arguments.update({arg: bool(arguments.get(arg)) for arg in boolean_args})

    return arguments


# def process_filter_args():
#     red_entry_day_exit = True if arguments["red_day_exit"] else False
#     failsafe = True if arguments["failsafe"] else False
#     return red_entry_day_exit, failsafe


def p2f(s):
    try:
        stripped_s = s.strip("%")
        if stripped_s == "":
            return
        else:
            return float(stripped_s) / 100
    except AttributeError:
        return s


def data_filter_by_dates(ws, start_date, end_date):
    ws = ws.loc[ws["entry_date"] >= start_date]
    ws = ws.loc[ws["entry_date"] <= end_date]
    return ws


def prepare_data(ws):
    # Convert types
    # This should be in gsheetobj
    num_cols = [
        "entry_price_planned",
        "entry_price_actual",
        "exit_price_planned",
        "main_exit_price",
        "threshold_1_expected_price",
        "threshold_1_actual_price",
        "threshold_2_expected_price",
        "threshold_2_actual_price",
        "threshold_3_expected_price",
        "threshold_3_actual_price",
        "weekly_mri_count",
        "fisher_daily",
        "fisher_weekly",
        "coppock_daily",
        "coppock_weekly"
    ]
    ws[num_cols] = ws[num_cols].apply(pd.to_numeric, errors="coerce")

    ws["max_level_reached"] = ws["max_level_reached"].apply(p2f)
    ws["entry_date"] = pd.to_datetime(
        ws["entry_date"], format="%d/%m/%Y", errors="coerce"
    )
    ws["control_exit_date"] = pd.to_datetime(
        ws["control_exit_date"], format="%d/%m/%Y", errors="coerce"
    )

    # Not needed in the new format
    for column in [
        "control_result_%",
        "exit_price_portion",
        "threshold_1_exit_portion",
        "threshold_2_exit_portion",
        "threshold_3_exit_portion",
        "max_level_reached",
    ]:
        ws[column] = ws[column].apply(p2f)

    return ws


def process_entry(sim, stock, entry_price, take_profit_variant):
    sim.set_take_profit_levels(stock, take_profit_variant)

    if len(sim.current_positions) + 1 > current_simultaneous_positions:
        print(f"max possible positions held, skipping {stock}")
    else:
        sim.positions_held += 1
        sim.current_positions.add(stock)
        sim.capital_per_position[stock] = sim.current_capital / current_simultaneous_positions
        sim.entry_prices[stock] = entry_price
        sim.set_take_profit_levels(stock, take_profit_variant)

        print(f"-> ENTER {stock} | positions held {sim.positions_held}")
        print(f'accounting for the brokerage: ${config["simulator"]["commission"]}')
        sim.current_capital -= config["simulator"]["commission"]
        print(f"current capital on entry: ${sim.current_capital}, allocated to the position: ${sim.capital_per_position[stock]}")


def exit_all_positions(sim, current_date_dt):
    for stock in list(sim.current_positions):  # Use list() to avoid modifying set during iteration
        price_data = get_price_from_db(stock, current_date_dt)
        if price_data:
            process_exit(sim, stock, price_data)
        else:
            print(f"Warning: No price data found for {stock} on {current_date_dt}. Skipping exit.")


def process_exit(sim, stock_code, price_data, partial=False):
    if stock_code in sim.current_positions:
        if partial:
            # Handle partial exit
            exit_proportion = sim.take_profit_info[stock_code]['total_exit_proportion']
            total_profit = sim.take_profit_info[stock_code]['total_profit']
        else:
            # Full exit
            exit_proportion = 1.0
            total_profit = (price_data['open'] - sim.entry_prices[stock_code]) / sim.entry_prices[stock_code]  # use open price on the exit day

        position_size = sim.capital_per_position[stock_code]
        exit_size = position_size * exit_proportion
        capital_gain = exit_size * total_profit

        sim.current_capital += capital_gain
        sim.current_capital -= config["simulator"]["commission"]

        print(f"-> EXIT {stock_code}: {'partial' if partial else 'full'} | proportion {exit_proportion:.2%} | Result: {total_profit:.2%} | Positions held: {sim.positions_held}")
        print(f"-> Position size at the exit: {position_size}")

        if not partial:
            print(f"-> Price at the full exit: {price_data['open']} | entry {sim.entry_prices[stock_code]}")
            sim.current_positions.remove(stock_code)
            sim.positions_held -= 1
            del sim.capital_per_position[stock_code]
            del sim.take_profit_info[stock_code]
        else:
            sim.capital_per_position[stock_code] -= exit_size

        sim.update_capital(sim.current_capital)  # Update capital values
        print(f"Capital after exit: ${sim.current_capital}")

        # Update trade statistics
        sim.update_trade_statistics(total_profit)


def update_results_dict(
    results_dict,
    sim,
    current_simultaneous_positions,
    take_profit_variant_name,
    current_variant="control",
    extra_suffix="",
):
    result_current_dict = dict(
        growth=sim.growth,
        win_rate=sim.win_rate,
        winning_trades_number=sim.winning_trades_number,
        losing_trades_number=sim.losing_trades_number,
        best_trade_adjusted=sim.best_trade_adjusted,
        worst_trade_adjusted=sim.worst_trade_adjusted,
        max_drawdown=sim.max_drawdown,
        max_negative_strike=sim.max_negative_strike,
        median_mom_growth=sim.mom_growth,
        average_mom_growth=sim.average_mom_growth,
        simultaneous_positions=current_simultaneous_positions,
        variant_group=current_variant,
    )
    results_dict[
        f"{current_variant}_{current_simultaneous_positions}pos_{take_profit_variant_name}{extra_suffix}"
    ] = result_current_dict
    return results_dict


def add_entry_with_profit_thresholds(sim, stock, entry_price_actual, entry_date_actual):
    if len(sim.current_positions) + 1 > current_simultaneous_positions:
        print(f"max possible positions held, skipping {stock}")
    else:
        sim.positions_held += 1
        sim.current_positions.add(stock)
        sim.capital_per_position[stock] = (
            sim.current_capital / current_simultaneous_positions
        )

        print(
            "-> entry",
            stock,
            "| positions held",
            sim.positions_held,
        )
        print(f"-> entry price: {entry_price_actual}")
        print(f'accounting for the brokerage: ${config["simulator"]["commission"]}')
        sim.current_capital -= config["simulator"]["commission"]
        print(
            f"current_capital: ${sim.current_capital}, allocated to the position: ${sim.capital_per_position[stock]}"
        )

        # on the entry, we have the full position
        sim.left_of_initial_entries[stock] = 1
        sim.entry_prices[stock] = entry_price_actual
        sim.entry_dates[stock] = entry_date_actual

        # also, on the entry we initiate the dict of thresholds hit for the item
        # they will then be populated with like (0.25, ...)
        sim.thresholds_reached[
            stock
        ] = set()  # appropriate as each would only be there once


def failsafe_trigger_check(sim, stock_prices, current_date_dt):
    for position in sim.current_positions:
        if position not in sim.failsafe_stock_trigger:
            current_df = stock_prices[position][0]
            curr_row = current_df.loc[current_df["timestamp"] == current_date_dt]
            failsafe_current_level = sim.entry_prices[position] * (1 + config["simulator"]["failsafe_trigger_level"])
            if not curr_row.empty:
                if curr_row["high"].iloc[0] >= failsafe_current_level:
                    print(f"failsafe level reached for {position} @ {failsafe_current_level}")
                    sim.failsafe_stock_trigger[position] = True
                    sim.failsafe_active_dates[position] = current_date_dt


def failsafe_trigger_rollback(sim, stock_prices, current_date_dt):
    failback_triggers = []
    for position in sim.current_positions:
        current_df = stock_prices[position][0].copy()
        current_df['next_open'] = current_df['open'].shift(-1)

        curr_row = current_df.loc[current_df["timestamp"] == current_date_dt]

        failsafe_rollback_level = sim.entry_prices[position] * (1 + config["simulator"]["failsafe_exit_level"])
        if not curr_row.empty and (position in sim.failsafe_stock_trigger):
            if sim.failsafe_stock_trigger[position]:
                print(f'{position} failsafe levels check: curr_low {curr_row["low"].iloc[0]} | fsafe level: {failsafe_rollback_level} | failsafe_date {sim.failsafe_active_dates[position]}')
                if (curr_row["low"].iloc[0] < failsafe_rollback_level) and (sim.failsafe_active_dates[position] != current_date_dt):
                    print(f"failsafe rollback for {position} @ {failsafe_rollback_level} on {current_date_dt}")

                    # We should use the correct price
                    #price_to_use = min(curr_row["open"].iloc[0], failsafe_rollback_level) # old incorrect logic
                    price_to_use = curr_row["next_open"].iloc[0]
                    print(f'-- using the price {price_to_use}: next day open')

                    failback_triggers.append([position, price_to_use])

    return failback_triggers


def thresholds_check(sim, stock_prices, current_date_dt):
    for position in sim.current_positions:
        current_df = stock_prices[position][0]
        curr_row = current_df.loc[current_df["timestamp"] == current_date_dt]
        if not curr_row.empty:
            print(
                f"current high for {position} {curr_row['high'].iloc[0]} | entry: {sim.entry_prices[position]}"
            )
            for each_threshold in current_tp_variant:
                if curr_row["high"].iloc[0] > sim.entry_prices[position] * (
                    1 + each_threshold
                ):
                    # Decrease the residual
                    if each_threshold not in sim.thresholds_reached[position]:
                        sim.left_of_initial_entries[position] -= (
                            1 / current_simultaneous_positions
                        )
                    # Add the threshold
                    sim.thresholds_reached[position].add(each_threshold)
                    print(f"-- {position} reached {each_threshold:.2%}")


def add_exit_with_profit_thresholds(
    sim, stock, entry_price_actual, exit_price_actual, control_result_percent
):
    if stock in sim.current_positions:
        position = stock
        sim.current_positions.remove(stock)
        sim.positions_held -= 1

        # check what thresholds were reached. use entry and exit price and thresholds reached
        number_thresholds_total = len(current_tp_variant)
        number_thresholds_reached = len(sim.thresholds_reached[position])
        divisor = number_thresholds_total + 1
        portion_not_from_thresholds = divisor - number_thresholds_reached

        exit_price_in_calc = exit_price_actual
        print(
            f"exit price used for {position}: {exit_price_in_calc}, entry price: {entry_price_actual}"
        )

        absolute_price_result = (
            exit_price_in_calc - entry_price_actual
        ) / entry_price_actual
        result = absolute_price_result * portion_not_from_thresholds / divisor
        print(
            f"absolute price change result for {position}: {absolute_price_result:.2%} | "
            f"multiplier (considering thresholds): {portion_not_from_thresholds}/{divisor}"
        )
        print(f"relative price change result for {position}: {result:.2%}")
        print(
            f"thresholds reached ({position}): {sim.thresholds_reached[position]}: {number_thresholds_reached} of {number_thresholds_total}"
        )

        # Correct brokerage
        number_brokerage_commissions_paid = number_thresholds_reached + 1

        for threshold_reached in sim.thresholds_reached[position]:
            print(f"--- calc: extra result += {threshold_reached}/{divisor}")
            result += threshold_reached / divisor

        print(f"result ({position}) accounting for thresholds reached: {result:.2%}")

        if result >= 0:
            sim.winning_trades_number += 1
            sim.winning_trades.append(result)
            if (sim.best_trade_adjusted is None) or (
                result / current_simultaneous_positions > sim.best_trade_adjusted
            ):
                sim.best_trade_adjusted = result / current_simultaneous_positions
                print(f"best_trade_adjusted is now {sim.best_trade_adjusted}")
        elif result < 0:
            sim.losing_trades_number += 1
            sim.losing_trades.append(result)
            if (sim.worst_trade_adjusted is None) or (
                result / current_simultaneous_positions < sim.worst_trade_adjusted
            ):
                sim.worst_trade_adjusted = result / current_simultaneous_positions

        # Add to all trades
        sim.all_trades.append(result)

        print(
            f"-> exit {stock} | result: {result:.2%} | positions held {sim.positions_held}"
        )

        position_size = sim.capital_per_position[stock]
        print(f"allocated to the position originally: ${position_size}")
        capital_gain = position_size * result
        print(f"capital gain/loss: ${capital_gain}".replace("$-", "-$"))
        print(f"capital state pre exit: ${sim.current_capital}")

        sim.current_capital = sim.current_capital + capital_gain
        print(
            f'accounting for the brokerage: ${config["simulator"]["commission"] * number_brokerage_commissions_paid} ({config["simulator"]["commission"]}x{number_brokerage_commissions_paid})'
        )
        sim.current_capital -= config["simulator"]["commission"] * number_brokerage_commissions_paid
        print(f"balance: ${sim.current_capital}")
        sim.capital_values.append(sim.current_capital)

        # Delete from the partial positions left, prices, thresholds for the element
        sim.remove_stock_traces(stock)


# plotting
def plot_latest_sim(latest_sim):
    x, y = [], []
    for key, value in latest_sim.detailed_capital_values.items():
        x.append(key)
        y.append(value)
    _ = plt.plot(x, y)
    ax = plt.gca()
    plt.xticks(fontsize=7)
    lst = list(range(1000))
    lst = lst[0::20]
    for index, label in enumerate(ax.get_xaxis().get_ticklabels()):
        if index not in lst:
            label.set_visible(False)
    plt.show()


def failed_entry_day_check(sim, stock_prices, stock_name, current_date_dt):
    if len(sim.current_positions) + 1 > current_simultaneous_positions:
        print(f"max possible positions held, skipping {stock_name}")
    else:
        if red_entry_day_exit:
            stock_prices_df = stock_prices[stock_name][0]
            stock_volume_df = stock_prices[stock_name][1]
            curr_state_price = stock_prices_df.loc[
                stock_prices_df["timestamp"] <= current_date_dt
            ]
            curr_state_volume = stock_volume_df.loc[
                stock_volume_df["timestamp"] <= current_date_dt
            ]
            warning, _ = red_day_on_volume(
                curr_state_price, curr_state_volume, output=True, stock_name=stock_name
            )
            if warning:
                sim.failed_entry_day_stocks[stock_name] = True


def failed_entry_day_process(sim, stock_prices, current_date_dt):
    failed_entry_day_stocks_to_iterate = sim.failed_entry_day_stocks.copy()
    for stock_name, elem in failed_entry_day_stocks_to_iterate.items():
        print(f"Failed entry day for {stock_name}, exiting")
        current_df = stock_prices[stock_name][0]
        curr_row = current_df.loc[current_df["timestamp"] == current_date_dt]
        if not curr_row.empty:
            print(f"Current open for {stock_name}: {curr_row['open'].iloc[0]}")
            add_exit_with_profit_thresholds(
                sim,
                stock_name,
                sim.entry_prices[stock_name],
                curr_row["open"].iloc[0],
                None,
            )


def get_dates(start_date, end_date):
    try:
        start_date_dt = datetime.strptime(start_date, "%Y-%m-%d")
        end_date_dt = datetime.strptime(end_date, "%Y-%m-%d")
    except ValueError:
        print("the date must be in the format YYYY-MM-DD")
        exit(0)
    current_date_dt = start_date_dt
    return start_date_dt, end_date_dt, current_date_dt


def check_profit_levels(sim, current_positions, current_date_dt, take_profit_variant):
    for stock in current_positions:
        price_data = get_price_from_db(stock, current_date_dt)
        if price_data:
            if sim.check_and_update_take_profit(stock, price_data['high'], price_data['open'], take_profit_variant):
                # Process partial exit
                process_exit(sim, stock, price_data, partial=True)  # take_profit_variant['take_profit_values']

def run_simulation(results_dict, take_profit_variant):
    sim = Simulation(capital=config["simulator"]["capital"])
    print(f"Take profit variant {take_profit_variant['variant_name']} | max positions {current_simultaneous_positions}")

    start_date_dt, end_date_dt, current_date_dt = get_dates(start_date, end_date)

    sim.balances[start_date_dt.strftime("%d/%m/%Y")] = sim.current_capital
    sim.detailed_capital_values[start_date_dt.strftime("%d/%m/%Y")] = sim.current_capital

    while current_date_dt < end_date_dt:
        previous_date_month = current_date_dt.strftime("%m")
        current_date_dt = current_date_dt + timedelta(days=1)
        current_date_month = current_date_dt.strftime("%m")

        if previous_date_month != current_date_month:
            sim.balances[current_date_dt.strftime("%d/%m/%Y")] = sim.current_capital
        sim.detailed_capital_values[current_date_dt.strftime("%d/%m/%Y")] = sim.current_capital

        print(current_date_dt, "| positions: ", sim.current_positions)

        # Entries
        day_entries = ws.loc[ws["entry_date"] == current_date_dt]
        for key, row in day_entries.iterrows():
            process_entry(sim, row["stock"], row["entry_price_actual"], take_profit_variant)

        # Check whether stocks reach the profit level for each stock
        if len(sim.current_positions) > 0:
            check_profit_levels(sim, sim.current_positions, current_date_dt, take_profit_variant)

        # Exits
        day_exits = ws.loc[ws[f"control_exit_date"] == current_date_dt]
        for key, row in day_exits.iterrows():
            price_data = get_price_from_db(row["stock"], current_date_dt)
            if price_data:
                process_exit(sim, row["stock"], price_data)

        sim.update_capital(sim.current_capital)  # Update capital values at the end of each day

    # Exit all remaining positions at the end of the simulation
    if len(sim.current_positions) > 0:
        print(f"[x] Stopped similation: exiting all remaining positions as of {end_date_dt}")
        exit_all_positions(sim, end_date_dt)

    # Add one more entry for the start of the next month
    next_month = (end_date_dt.replace(day=1) + timedelta(days=32)).replace(day=1)
    sim.balances[next_month.strftime("%d/%m/%Y")] = sim.current_capital

    # Calculate metrics and print the results
    sim.calculate_metrics()
    sim.print_metrics()

    # Saving the result in the overall dictionary
    results_dict = update_results_dict(
        results_dict, sim, current_simultaneous_positions, take_profit_variant['variant_name']
    )
    return results_dict, sim

def create_monthly_breakdown(simulations):
    monthly_data = {}
    for variant, sim in simulations.items():
        for date, value in sim.balances.items():
            date_obj = datetime.strptime(date, "%d/%m/%Y")
            month_start = date_obj.strftime("%d/%m/%Y")  # Keep the day as-is
            if month_start not in monthly_data:
                monthly_data[month_start] = {}
            monthly_data[month_start][variant] = value

    df = pd.DataFrame(monthly_data).T
    df.index.name = 'Date'
    df.reset_index(inplace=True)
    return df

# def show_monthly_breakdown(result, positions):
#     # Open a file for writing
#     csv_filename = f'sim_monthly.csv'
#
#     with open(csv_filename, 'w', newline='') as csv_file:
#         csv_writer = csv.writer(csv_file)
#
#         # Write the header
#         csv_writer.writerow(['Date', 'Value'])
#
#         # Convert string dates to datetime objects
#         date_values = {datetime.strptime(date, '%d/%m/%Y'): value for date, value in result.items()}
#
#         print()
#         print('(!) Note: monthly results are only shown for the last simulation')
#         print('--- Monthly breakdown ---')
#
#         # Iterate over the dictionary and print values at the beginning of each month
#         current_month = None
#         for date, value in sorted(date_values.items()):
#             if date.month != current_month:
#                 current_month = date.month
#                 formatted_date = date.replace(day=1).strftime("%d/%m/%Y")
#                 rounded_value = round(value, 1)
#                 print(f'{formatted_date}: {rounded_value}')
#                 csv_writer.writerow([formatted_date, rounded_value])
#
#         print('-------------------------')
#         print(f'(i) results have been written to {csv_filename}')
#         print()


# Apply filters from config to the DataFrame
def filter_dataframe(df, config):
    filters = config["simulator"]["numerical_filters"]
    for column, conditions in filters.items():
        if isinstance(conditions, dict):  # Ensure conditions is a dictionary
            if 'min' in conditions:
                df = df[df[column] >= conditions['min']]
            if 'max' in conditions:
                df = df[df[column] <= conditions['max']]
        else:
            print(f"Error: Filter conditions for {column} are not specified correctly.")
    return df


# Get and save stock prices if not available for running the simulation
def get_stock_prices(sheet_df, prices_start_date):
    stock_names = [item.stock for key, item in sheet_df.iterrows()]
    stock_markets = [item.market for key, item in sheet_df.iterrows()]

    # Convert prices_start_date to date if it's not already
    if isinstance(prices_start_date, str):
        prices_start_date = datetime.strptime(prices_start_date, '%Y-%m-%d').date()
    elif isinstance(prices_start_date, datetime):
        prices_start_date = prices_start_date.date()

    # Check if the Price table exists and get the earliest date
    earliest_date = check_earliest_price_date()

    # Ensure earliest_date is also a date object for comparison
    if earliest_date:
        if isinstance(earliest_date, datetime):
            earliest_date = earliest_date.date()
        elif isinstance(earliest_date, str):
            earliest_date = datetime.strptime(earliest_date, '%Y-%m-%d').date()

        # Check if the earliest date in the database is within 5 days of the start date
        date_difference = abs((earliest_date - prices_start_date).days)
        if (date_difference <= 5) and not arguments["forced_price_update"]:
            print(f"Data within 5 days of {prices_start_date} already exists in the database. Skipping update.")
            return {}

    # If the dates don't match or the table was just created, proceed with updating
    delete_all_prices()  # Clear existing data

    stock_prices = {}
    prices_to_add = []
    added_records = set()  # To keep track of unique (stock, date) combinations

    for i, stock in enumerate(stock_names):
        market = Market(stock_markets[i])
        print(f"Getting stock data for {stock}{market.stock_suffix}")
        stock_df = get_stock_data(f"{stock}{market.stock_suffix}", prices_start_date)
        stock_df = stock_df[0]  # this is the actual dataframe

        stock_prices[stock] = stock_df.to_dict('records')

        # Prepare data for bulk insert, avoiding duplicates
        for _, row in stock_df.iterrows():
            record_key = (stock, row['timestamp'])
            if record_key not in added_records:
                prices_to_add.append({
                    'stock': stock,
                    'date': row['timestamp'].to_pydatetime(),
                    'open': row['open'],
                    'high': row['high'],
                    'low': row['low'],
                    'close': row['close']
                })
                added_records.add(record_key)

    # Bulk insert the new price data
    if prices_to_add:
        try:
            bulk_add_prices(prices_to_add)
        except IntegrityError as e:
            print(f"Error inserting prices: {e}")
            print("Some records may already exist in the database. Continuing with available data.")

    return stock_prices


############## MAIN CODE ###############
if __name__ == "__main__":

    arguments = define_args()
    #red_entry_day_exit, failsafe = process_filter_args()

    print("reading the values...")

    # Dates for simulation
    start_date = arguments["start"]
    end_date = arguments["end"]

    # Price data start date is required in advance because we are calculating on a weekly scale
    prices_start_date = datetime.strptime(start_date, "%Y-%m-%d") - timedelta(
        days=2 * 365
    )
    # ^^^ -2 years ago from start is ok for data handling
    prices_start_date = prices_start_date.strftime("%Y-%m-%d")

    # Get the sheet to a dataframe
    sheet_name = config["logging"]["gsheet_name"]
    tab_name = config["logging"]["gsheet_tab_name"]

    ws = gsheetsobj.sheet_to_df(sheet_name, tab_name)

    ws.columns = config["logging"]["gsheet_columns"]
    ws = prepare_data(ws)

    # Dict to save all the results, start with empty
    results_dict = dict()

    # Dates filtering for the dataset in the sheet
    start_date_dt, end_date_dt, current_date_dt = get_dates(start_date, end_date)
    ws = data_filter_by_dates(ws, start_date_dt, end_date_dt)

    # TEST TEST
    #ws = ws[ws['stock'] == 'AGIO']

    # Filter the dataset per the config for the numerical parameters
    ws = filter_dataframe(ws, config)

    # Get information on the price data if the date is new
    get_stock_prices(ws, prices_start_date)

    # > Iterate over take profit variants and number of positions
    simulations = {}

    for current_simultaneous_positions in config["simulator"]["simultaneous_positions"]:
        for take_profit_variant in config["simulator"]["take_profit_variants"]:
            # For logging
            variant_name = f"{current_simultaneous_positions}pos_{take_profit_variant['variant_name']}"
            # Run it
            results_dict, latest_sim = run_simulation(
                results_dict, take_profit_variant
            )
            simulations[variant_name] = latest_sim

            # # Show breakdown
            # if arguments["show_monthly"]:
            #     show_monthly_breakdown(latest_sim.detailed_capital_values,
            #                            positions=config["simulator"]["simultaneous_positions"])

    # < Stop iterations

    # Finalisation

    # Write the output to a dataframe and a spreadsheet
    resulting_dataframes = []
    for k, v in results_dict.items():
        print(k, v)
        values_current = v.copy()
        values_current["variant"] = k
        resulting_dataframes.append(
            pd.DataFrame.from_records(values_current, index=[0])
        )

    # final_result = pd.concat(df for df in resulting_dataframes)
    # final_result = final_result[
    #     [
    #         "variant",
    #         "simultaneous_positions",
    #         "variant_group",
    #         "best_trade_adjusted",
    #         "growth",
    #         "losing_trades_number",
    #         "max_drawdown",
    #         "max_negative_strike",
    #         "win_rate",
    #         "median_mom_growth",
    #         "average_mom_growth",
    #         "winning_trades_number",
    #         "worst_trade_adjusted",
    #     ]
    # ]
    #
    # # save to csv
    # final_result.to_csv("sim_summary.csv", index=False)
    # print()
    # print("(i) summary saved to sim_summary.csv")

    # Create the summary DataFrame
    final_result = pd.concat(pd.DataFrame.from_records(v, index=[0]) for v in results_dict.values())
    final_result['variant'] = results_dict.keys()
    final_result = final_result[
        [
            "variant",
            "simultaneous_positions",
            "variant_group",
            "best_trade_adjusted",
            "growth",
            "losing_trades_number",
            "max_drawdown",
            "max_negative_strike",
            "win_rate",
            "median_mom_growth",
            "average_mom_growth",
            "winning_trades_number",
            "worst_trade_adjusted",
        ]
    ]

    # Format percentage and number columns
    percentage_cols = ["best_trade_adjusted", "growth", "max_drawdown", "win_rate", "median_mom_growth", "average_mom_growth", "worst_trade_adjusted"]
    number_cols = ["losing_trades_number", "max_negative_strike", "winning_trades_number"]

    final_result[percentage_cols] = final_result[percentage_cols].applymap(format_percentage)
    final_result[number_cols] = final_result[number_cols].applymap(format_number)

    # Create the monthly breakdown DataFrame
    monthly_breakdown = create_monthly_breakdown(simulations)

    # Create the monthly breakdown DataFrame
    monthly_breakdown = create_monthly_breakdown(simulations)
    monthly_breakdown['Date'] = pd.to_datetime(monthly_breakdown['Date'], format='%d/%m/%Y')
    monthly_breakdown = monthly_breakdown.sort_values('Date')
    monthly_breakdown['Date'] = monthly_breakdown['Date'].dt.strftime('%d/%m/%Y')

    # Format monthly breakdown values
    monthly_breakdown.iloc[:, 1:] = monthly_breakdown.iloc[:, 1:].applymap(format_number)


    # Create Excel file with three sheets
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "Summary"

    for r in dataframe_to_rows(final_result, index=False, header=True):
        ws1.append(r)

    ws2 = wb.create_sheet(title="Monthly Breakdown")
    for r in dataframe_to_rows(monthly_breakdown, index=False, header=True):
        ws2.append(r)

    # Adjust column width and set font size for both sheets
    for ws in [ws1, ws2]:
        adjust_column_width(ws)
        set_font_size(ws, 12)

    # Add plots if the plot argument is provided
    if arguments["plot"]:
        ws3 = wb.create_sheet(title="Plots")
        row = 1
        for variant_name, sim in simulations.items():
            img_buf = create_variant_plot(sim, variant_name)
            img = Image(img_buf)
            img.width = 900
            img.height = 500
            ws3.add_image(img, f'A{row}')

            row += 30  # Spacing between plots

    # Save the Excel file
    excel_filename = "sim_summary.xlsx"
    wb.save(excel_filename)
    print(f"(i) Detailed info saved to {excel_filename}")
