import argparse
import arrow
import pandas as pd

from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, numbers
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment

import matplotlib.pyplot as plt
import matplotlib.dates as mdates
from openpyxl.drawing.image import Image
import io

from datetime import datetime, timedelta


parser = argparse.ArgumentParser()


def get_data_start_date(input_date=None):
    if input_date is None:
        current_date = arrow.now()
    else:
        current_date = arrow.get(input_date.strftime("%Y-%m-%d"), "YYYY-MM-DD")

    shifted_date = current_date.shift(months=-12)
    data_start_date = shifted_date.format("YYYY-MM-DD")

    return data_start_date


def get_previous_workday():
    current_datetime = arrow.now()
    current_dow = current_datetime.isoweekday()
    if current_dow == 1:  # only subtract if today is Monday
        current_datetime = current_datetime.shift(days=-3)
    else:
        current_datetime = current_datetime.shift(days=-1)
    current_datetime = current_datetime.format("YYYY-MM-DD")
    return current_datetime


def get_current_workday():
    current_datetime = arrow.now()
    current_datetime = current_datetime.format("YYYY-MM-DD")
    return current_datetime


def define_args_method_only():
    parser.add_argument(
        "-method",
        type=str,
        required=True,
        choices=["mri", "anx"],
        help="Method (mri or anx)"
    )

    args = parser.parse_args()
    arguments = vars(args)

    return arguments

def define_simulator_args():
    # Take profit levels variation is only supported for the control group, thus the modes are different
    # Removing this as not used now, only one mode using the sheet
    # parser.add_argument(
    #     "-mode",
    #     type=str,
    #     required=True,
    #     help="Mode to run the simulation in (main|tp). Main mode means taking profit as per the spreadsheet setup.",
    #     choices=["main", "tp"],
    # )
    parser.add_argument(
        "--plot", action="store_true", help="Plot the latest simulation"
    )
    parser.add_argument(
        "--failsafe", action="store_true", help="Activate the failsafe approach"
    )
    parser.add_argument(
        "--forced_price_update", action="store_true", help="Activate the failsafe approach"
    )
    #
    # parser.add_argument(
    #     "--show_monthly", action="store_true", help="Show MoM capital value (only in main mode)"
    # )

    parser.add_argument(
        "-method",
        type=str,
        required=True,
        choices=["mri", "anx"],
        help="Method of shortlisting (mri or anx)"
    )

    # To run on one stock
    parser.add_argument(
        "-stock",
        type=str,
        required=False,
        help="Specify a single stock to run the simulation on"
    )

    # Adding the dates
    parser.add_argument(
        "-start",
        type=str,
        required=True,
        help="Start date to run for (YYYY-MM-DD format)",
    )
    parser.add_argument(
        "-end",
        type=str,
        required=True,
        help="End date to run for (YYYY-MM-DD format)",
    )

    # To add averaged sampling
    parser.add_argument("--sampling", action="store_true", help="Enable sampling mode")

    # Not used
    # # Arguments to overwrite default settings for filtering
    # parser.add_argument(
    #     "--red_day_exit",
    #     action="store_true",
    #     help="Exit when entry day is red (in tp mode only)",
    # )

    # Not used
    # Test for the exit price variation experimentation
    # parser.add_argument(
    #     "--exit_variation_a",
    #     action="store_true",
    #     help="Exit approach experiment A (main mode only)",
    # )

    args = parser.parse_args()
    arguments = vars(args)

    # Convert specific arguments to boolean, defaulting to False if not provided
    boolean_args = ["plot", "failsafe", "forced_price_update", "sampling"]   # "show_monthly"
    arguments.update({arg: bool(arguments.get(arg)) for arg in boolean_args})

    # Convert stock to upper case
    if arguments['stock'] is not None:
        arguments['stock'] = arguments['stock'].upper()

    return arguments


# def process_filter_args():
#     red_entry_day_exit = True if arguments["red_day_exit"] else False
#     failsafe = True if arguments["failsafe"] else False
#     return red_entry_day_exit, failsafe


def define_scanner_args():
    parser.add_argument(
        "--update",
        action="store_true",
        help="Update the assets list. Do this before scanning.",
    )
    parser.add_argument(
        "--scan", action="store_true", help="Scan for potential signals"
    )
    parser.add_argument(
        "-date",
        type=str,
        required=False,
        help="Date to run as of (YYYY-MM-DD format) for update or scan",
    )
    parser.add_argument(
        "-num", type=int, required=False, help="Limit the number of scanned stocks"
    )
    parser.add_argument(
        "-method",
        type=str,
        required=False,
        choices=["mri", "anx"],
        help="Method of shortlisting (mri or anx)"
    )
    parser.add_argument(
        "-stocks", type=str, required=False, help="Force checking specific stocks only"
    )

    args = parser.parse_args()
    arguments = vars(args)

    if not arguments["update"]:
        arguments["update"] = False
    if not arguments["scan"]:
        arguments["scan"] = False
    if arguments["stocks"] is not None:
        arguments["stocks"] = arguments["stocks"].upper()

    # Process the date
    if arguments["date"] is not None:
        try:
            arguments["date"] = arrow.get(arguments["date"], "YYYY-MM-DD").naive
        except arrow.parser.ParserMatchError:
            print("The date must be in the format YYYY-MM-DD")
            exit(0)

    # Check if method is specified
    if arguments["scan"] and arguments["method"] is None:
            print("Specify the method when scanning")
            exit(0)

    if True not in arguments.values():
        print("No arguments specified. Run scanner.py --h to show help.")
        exit(0)

    return arguments


def dates_diff(date_from, date_to=None):
    date_to = arrow.now() if date_to is None else date_to
    return (date_to - date_from).days


def format_number(num):
    magnitude = 0
    while abs(num) >= 1000:
        magnitude += 1
        num /= 1000.0
    # add more suffixes if you need them
    return "%.2f%s" % (num, ["", "K", "M", "G", "T", "P"][magnitude])


def format_bool(value):
    # Format boolean as tick or fail
    formatted_value = "v" if value else "x"
    return formatted_value


def get_test_stocks():
    # Use for testing / bugfixes
    # In scanner.py, use: stocks = get_test_stocks()

    class Stk:
        code, name = None, None

    test_stock = Stk()
    test_stock.code = "NVDA"
    test_stock.name = "NVDA"
    return [test_stock]

############################################################################
############# Functions to create a simulation report #####################
############################################################################

def create_monthly_breakdown(simulations):
    monthly_data = {}
    for variant, sim in simulations.items():
        for date, value in sim.balances.items():
            date_obj = datetime.strptime(date, "%d/%m/%Y")
            month_start = date_obj.strftime("%d/%m/%Y")  # Keep the day as-is
            if month_start not in monthly_data:
                monthly_data[month_start] = {}
            monthly_data[month_start][variant] = value

    df = pd.DataFrame(monthly_data).T
    df.index.name = 'Date'
    df.reset_index(inplace=True)
    return df

def format_percentage(value):
    return value  # Return the raw value instead of formatted string

def format_number(value):
    return value  # Return the raw value instead of formatted string

def create_variant_plot(sim, variant_name):
    # Convert dates to datetime objects
    dates = [datetime.strptime(date, "%d/%m/%Y") for date in sim.detailed_capital_values.keys()]
    values = list(sim.detailed_capital_values.values())

    # Add termination value as a step change
    last_date = dates[-1]
    next_month = (last_date.replace(day=1) + timedelta(days=32)).replace(day=1)

    plt.figure(figsize=(12, 6))

    # Plot the main line
    plt.step(dates, values, where='post')

    # Add the final step
    plt.step([last_date, next_month], [values[-1], sim.current_capital], where='post')

    plt.title(f"Capital Over Time - {variant_name}")
    plt.xlabel("Date")
    plt.ylabel("Capital")

    # Set x-axis to show only the first of each month
    plt.gca().xaxis.set_major_locator(mdates.MonthLocator())
    plt.gca().xaxis.set_major_formatter(mdates.DateFormatter('%Y-%m-%01'))

    # Extend x-axis slightly to show the final step
    plt.xlim(dates[0], next_month + timedelta(days=1))

    plt.xticks(rotation=45)
    plt.tight_layout()

    # Save plot to a bytes buffer
    buf = io.BytesIO()
    plt.savefig(buf, format='png', dpi=300)
    buf.seek(0)
    plt.close()
    return buf

def adjust_column_width(worksheet):
    for column in worksheet.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = max_length + 2
        worksheet.column_dimensions[column_letter].width = adjusted_width

def set_font_size_and_alignment(worksheet, size):
    for row in worksheet.iter_rows():
        for cell in row:
            if cell.row == 1:  # Header row
                cell.font = Font(size=size, bold=True)
                cell.value = cell.value.replace('_', ' ').title()
            else:
                cell.font = Font(size=size)
            cell.alignment = Alignment(horizontal='center')

def set_font_size(worksheet, size):
    for row in worksheet.iter_rows():
        for cell in row:
            if cell.row == 1:  # Header row
                cell.font = Font(size=size, bold=True)
            else:
                cell.font = Font(size=size)

def print_metrics(results_dict):
    print("\n\n__________ SUMMARIES ____________")
    for k, v in results_dict.items():
        print(f"\n{k}")
        for metric, value in v.items():
            if isinstance(value, float):
                if metric in ['growth', 'win_rate', 'best_trade_adjusted', 'worst_trade_adjusted', 'max_drawdown', 'median_mom_growth', 'average_mom_growth']:
                    print(f"- {metric}: {value:.2%}")
                else:
                    print(f"- {metric}: {value:.4f}")
            elif isinstance(value, int):
                print(f"- {metric}: {value}")
            # don't need non numerical
    print()

def create_report(results_dict, simulations, plot):
    # Write the output to a dataframe and a spreadsheet
    resulting_dataframes = []

    print_metrics(results_dict)

    for k, v in results_dict.items():

        values_current = v.copy()
        values_current["variant"] = k
        resulting_dataframes.append(
            pd.DataFrame.from_records(values_current, index=[0])
        )

    # Create the summary DataFrame
    final_result = pd.concat(pd.DataFrame.from_records(v, index=[0]) for v in results_dict.values())
    final_result['variant'] = results_dict.keys()
    final_result = final_result[
        [
            "variant",
            "max_positions",
            "growth",
            "max_drawdown",
            "win_rate",
            "median_mom_growth",
            "average_mom_growth",
            "winning_trades_number",
            "losing_trades_number",
            "max_negative_strike",
            "best_trade_adjusted",
            "worst_trade_adjusted",
        ]
    ]

    # Format percentage and number columns
    percentage_cols = ["growth", "max_drawdown", "win_rate", "median_mom_growth", "average_mom_growth", "best_trade_adjusted", "worst_trade_adjusted"]
    number_cols = ["max_positions", "winning_trades_number", "losing_trades_number", "max_negative_strike"]

    final_result[percentage_cols] = final_result[percentage_cols].applymap(format_percentage)
    final_result[number_cols] = final_result[number_cols].applymap(format_number)

    # Create the monthly breakdown DataFrame
    monthly_breakdown = create_monthly_breakdown(simulations)

    # Create the monthly breakdown DataFrame
    monthly_breakdown = create_monthly_breakdown(simulations)
    monthly_breakdown['Date'] = pd.to_datetime(monthly_breakdown['Date'], format='%d/%m/%Y')
    monthly_breakdown = monthly_breakdown.sort_values('Date')
    monthly_breakdown['Date'] = monthly_breakdown['Date'].dt.strftime('%d/%m/%Y')

    # Format monthly breakdown values
    monthly_breakdown.iloc[:, 1:] = monthly_breakdown.iloc[:, 1:].applymap(format_number)

    # Create Excel file with three sheets
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "Summary"

    # Get the column indices for percentage and number columns
    percentage_col_indices = [final_result.columns.get_loc(col) + 1 for col in percentage_cols]
    number_col_indices = [final_result.columns.get_loc(col) + 1 for col in number_cols]

    bold_font = Font(bold=True)

    for r_idx, row in enumerate(dataframe_to_rows(final_result, index=False, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            cell = ws1.cell(row=r_idx, column=c_idx, value=value)
            if r_idx == 1:  # Header row
                cell.font = bold_font
            elif c_idx in percentage_col_indices:
                cell.number_format = numbers.FORMAT_PERCENTAGE_00
            elif c_idx in number_cols:
                cell.number_format = numbers.FORMAT_NUMBER

    ws2 = wb.create_sheet(title="Monthly Breakdown")
    for r_idx, row in enumerate(dataframe_to_rows(monthly_breakdown, index=False, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            cell = ws2.cell(row=r_idx, column=c_idx, value=value)
            if r_idx == 1:  # Header row
                cell.font = bold_font
            elif c_idx > 1:  # Skip date column
                cell.number_format = numbers.FORMAT_NUMBER_00

    # Adjust column width and set font size for both sheets
    for ws in [ws1, ws2]:
        set_font_size_and_alignment(ws, 11)
        adjust_column_width(ws)

    # Add plots if the plot argument is provided
    if plot:
        ws3 = wb.create_sheet(title="Plots")
        row = 1
        for variant_name, sim in simulations.items():
            img_buf = create_variant_plot(sim, variant_name)
            img = Image(img_buf)
            img.width = 900
            img.height = 500
            ws3.add_image(img, f'A{row}')

            row += 30  # Spacing between plots

    # Save the Excel file
    excel_filename = "sim_summary.xlsx"
    wb.save(excel_filename)
    print(f"(i) Detailed info saved to {excel_filename}")


#### Simulation functions ####
def p2f(s):
    try:
        stripped_s = s.strip("%")
        if stripped_s == "":
            return
        else:
            return float(stripped_s) / 100
    except AttributeError:
        return s


def data_filter_by_dates(ws, start_date, end_date):
    ws = ws.loc[ws["entry_date"] >= start_date]
    ws = ws.loc[ws["entry_date"] <= end_date]
    return ws

def data_filter_from_date(ws, start_date):
    ws = ws.loc[ws["entry_date"] >= start_date]
    return ws

def prepare_data(ws):
    # Convert types
    # This should be in gsheetobj
    num_cols = [
        "entry_price_planned",
        "entry_price_actual",
        "exit_price_planned",
        "main_exit_price",
        "initial_stop_loss",
        "threshold_1_expected_price",
        "threshold_1_actual_price",
        "threshold_2_expected_price",
        "threshold_2_actual_price",
        "threshold_3_expected_price",
        "threshold_3_actual_price",
        "weekly_mri_count",
        "fisher_daily",
        "fisher_weekly",
        "coppock_daily",
        "coppock_weekly"
    ]
    ws[num_cols] = ws[num_cols].apply(pd.to_numeric, errors="coerce")

    ws["max_level_reached"] = ws["max_level_reached"].apply(p2f)
    ws["entry_date"] = pd.to_datetime(
        ws["entry_date"], format="%d/%m/%Y", errors="coerce"
    )
    ws["control_exit_date"] = pd.to_datetime(
        ws["control_exit_date"], format="%d/%m/%Y", errors="coerce"
    )

    # Extra fields
    ws["under_td_resistance"] = ws["under_td_resistance"].astype(bool)

    # Not needed in the new format
    for column in [
        "control_result_%",
        "exit_price_portion",
        "threshold_1_exit_portion",
        "threshold_2_exit_portion",
        "threshold_3_exit_portion",
        "max_level_reached",
    ]:
        ws[column] = ws[column].apply(p2f)

    return ws


# Apply filters from config to the DataFrame
def filter_dataframe(df, config):
    filters = config["simulator"]["numerical_filters"]
    for column, conditions in filters.items():
        if isinstance(conditions, dict):  # Ensure conditions is a dictionary
            if 'min' in conditions:
                df = df[df[column] >= conditions['min']]
            if 'max' in conditions:
                df = df[df[column] <= conditions['max']]
        else:
            print(f"Error: Filter conditions for {column} are not specified correctly.")
    return df
